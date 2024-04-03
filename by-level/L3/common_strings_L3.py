#L3, common strings
#Input Files: L3_class (ANCOMBC), level-3-R (taxa-summary)


import csv
import openpyxl
from openpyxl import load_workbook

#Import combined sheet from taxa-summary
L2_wb = openpyxl.load_workbook("by-level/level-3-R.xlsx")

combined_sheet_name = "combined"
combined_sheet = L2_wb[combined_sheet_name]

#Define and search for stats row
stats_row_index = None
for row_index, row in enumerate(combined_sheet.iter_rows(values_only=True)):
    if all(val in row for val in [' ', 'mean', 'mean', 'median', 'median', 'count', 'count', 'min', 'min', 'max', 'max', 'sd', 'sd']):
        stats_row_index = row_index
        break

if stats_row_index is not None:
    combined_sheet_data = []
    # Start collecting data after the stats row
    for row_index, row in enumerate(combined_sheet.iter_rows(min_row=stats_row_index + 1, values_only=True), start=stats_row_index + 1):
        # Stop when encountering a row with None values
        if all(val is None for val in row):
            break
        combined_sheet_data.append(row)

#Output: combined_summ_stats.csv
output_combined_csv = 'L3_combined_summ_stats.csv'

with open(output_combined_csv, 'w', newline='') as csvfile:
    writer = csv.writer(csvfile)
    #write data
    for row in combined_sheet_data:
        writer.writerow(row)

#Initialize taxons contained in taxa-summary (ts)
taxon_ts = []

with open ('L3_combined_summ_stats.csv', 'r', newline='') as csvfile:
    reader = csv.reader(csvfile)
    
    for row in reader:
        if row: #check if row is not empty
            taxon_ts.append(row[0])

#Remove empty values from taxon_ts
taxon_ts = taxon_ts[2:]

#remove everything before "p__"
for i in range(len(taxon_ts)):
        taxon_ts[i] = taxon_ts[i].split("p__")[1] 

#Extract taxon data from ANCOMBC
taxon_list = []

ANCOMBC_dict = {}

with open("by-level/L3_class.csv", newline="") as L2_phylum_ABC:
    # Create csv reader object
    reader = csv.DictReader(L2_phylum_ABC, delimiter=",")
    
    # Iterate over each row of csv file
    for row in reader:
        # Append 'taxon' column to the taxon list
        taxon_list.append(row['taxon'])
        
        # Remove 'taxon' column from the row and add it to ANCOMBC_dict
        row_data = {key: value for key, value in row.items() if key != 'taxon'}
        ANCOMBC_dict[row['taxon']] = row_data

#split row data to only include name
for i in range(len(taxon_list)):
    taxon_list[i] = taxon_list[i].split("p__")[1]

#find common features
def find_common_strings(list1, list2):
    set1 = set(list1)
    set2 = set(list2)

    common_strings = set1.intersection(set2)

    return list(common_strings)

common_taxons = find_common_strings(taxon_list, taxon_ts)
print("Common taxons:", common_taxons)


#output list to test file L3_common_taxons.txt
common_taxons_string = str(common_taxons)

with open("L3_common_taxons.txt", 'w') as file:
    file.write(common_taxons_string)