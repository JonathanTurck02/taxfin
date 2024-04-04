#L6
#author: Jonathan Turck
#this script extracts common strings from ancombc and taxa summary combined table outputs

import csv
import openpyxl
from openpyxl import load_workbook

L6_wb = openpyxl.load_workbook("ds3_PLE_vs_HC/taxa-summary/level-6-R.xlsx")

#initialize sheet and cell objects
#L2_sheet_obj = L2_wb.active
#L2_cell_obj = L2_sheet_obj.cell(row=1, column=1)

combined_sheet_name = "combined"
combined_sheet = L6_wb[combined_sheet_name]

# for row in combined_sheet.iter_rows(values_only=True):
#    print(row)

#write to csv file
csv_filename = "combined_sheet_data.csv"
with open(csv_filename, 'w', newline='') as csvfile:
    # Create a CSV writer object
    csvwriter = csv.writer(csvfile)
    
    # Iterate through each row in the combined sheet and write it to the CSV file
    for row in combined_sheet.iter_rows(values_only=True):
        csvwriter.writerow(row)

print(f"CSV file '{csv_filename}' has been created successfully.")

#' ', 'mean', 'mean', 'median', 'median', 'count', 'count', 'min', 'min', 'max', 'max', 'sd', 'sd'
#define header row in combined sheet and extract information

#Find the index of the row containing the header
stats_row_index = None
for row_index, row in enumerate(combined_sheet.iter_rows(values_only=True)):
    if all(val in row for val in [' ', 'mean', 'mean', 'median', 'median', 'count', 'count', 'min', 'min', 'max', 'max', 'sd', 'sd']):
        stats_row_index = row_index
        break

if stats_row_index is not None:
    combined_sheet_data = []
    # Start collecting data after the stats row
    for row_index, row in enumerate(combined_sheet.iter_rows(min_row=stats_row_index + 1, values_only=True), start=stats_row_index + 1):
        # Stop when encountering a row with None values
        if all(val is None for val in row):
            break
        combined_sheet_data.append(row)



#for row in combined_sheet_data:
#    print(row)

import csv

output_combined_csv = 'combined_summ_stats.csv'

with open(output_combined_csv, 'w', newline='') as csvfile:
    writer = csv.writer(csvfile)
    # Write the header
    #writer.writerow([' ', 'mean', 'mean', 'median', 'median', 'count', 'count', 'min', 'min', 'max', 'max', 'sd', 'sd'])
    # Write the data
    for row in combined_sheet_data:
        writer.writerow(row)

#initialize new taxon variable
taxon_ts = []

with open ('combined_summ_stats.csv', 'r', newline='') as csvfile:
    reader = csv.reader(csvfile)
    
    for row in reader:
        if row: #check if row is not empty
            taxon_ts.append(row[0])



#remove empty values from taxon_ts
taxon_ts = taxon_ts[2:]
#for i in taxon_ts:
    #print(i)

#remove everything before "p__"
for i in range(len(taxon_ts)):
        taxon_ts[i] = taxon_ts[i].split("g__")[1]  

#print(taxon_ts)  

#taxonomy information has been extracted from ts
#now extract taxonomy information from ANCOMBC

# Initialize an empty list to store the taxon data
taxon_list = []

# Initialize an empty dictionary to store the ANCOMBC data
ANCOMBC_dict = {}

with open("ds3_PLE_vs_HC/ANCOMBC/L6_genus.csv", newline="") as L6_class_ABC:
    # Create csv reader object
    reader = csv.DictReader(L6_class_ABC, delimiter=",")
    
    # Iterate over each row of csv file
    for row in reader:
        # Append 'taxon' column to the taxon list
        taxon_list.append(row['taxon'])
        
        # Remove 'taxon' column from the row and add it to ANCOMBC_dict
        row_data = {key: value for key, value in row.items() if key != 'taxon'}
        ANCOMBC_dict[row['taxon']] = row_data

#print(ANCOMBC_dict)
# Now taxon_list contains the taxon data and ANCOMBC_dict contains the ANCOMBC data
# print(taxon_list)
# split row data to only include name
for i in range(len(taxon_list)):
    taxon_list[i] = taxon_list[i].split("g__")[1]


#find common features
def find_common_strings(list1, list2):
    set1 = set(list1)
    set2 = set(list2)

    common_strings = set1.intersection(set2)

    return list(common_strings)

common_taxons = find_common_strings(taxon_list, taxon_ts)
print("Common taxons:", common_taxons)